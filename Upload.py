from MyProgram import Auth
import openpyxl as xl
import os

WorkBook = xl.load_workbook('StaffInfo/StaffInfo.xlsx')
Sheet = WorkBook['Sheet2']

def upload_image(client,config,image_path):

    print()
    print('Uploading....')
    image = client.upload_from_path(image_path,config=config,anon=False)

    return image


def imges():
    Register = False
    for imageFile in os.listdir('StaffCards'):
        ImageName, ImageEx = os.path.splitext(imageFile)
        StaffName, StaffRank = ImageName.split('_')

        for row in range(2, Sheet.max_row + 1):
            if Sheet.cell(row, 2).value == StaffName:
                if Sheet.cell(row, 7).value == 'Done':
                    Register = True

        if Register:
            print('{} is already Uploaded'.format(StaffName))
            continue

        config = {
            'album': None,
            'name': StaffName,
            'title': StaffRank,
            'description': 'None'
        }

        Image_Path = 'StaffCards/' + imageFile
        client = Auth.authenticate()
        image = upload_image(client, config, Image_Path)
        print('{} Card posed'.format(StaffName))
        print('YOu can get that here: {}'.format(image['link']))

        for row in range(2, Sheet.max_row + 1):
            if Sheet.cell(row, 2).value == StaffName:
                Sheet.cell(row, 7).value = 'Done'
                print('{} has been registed'.format(StaffName))
                WorkBook.save('StaffInfo/StaffInfo.xlsx')
        print('Done')




