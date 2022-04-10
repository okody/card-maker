import openpyxl as xl
from MyProgram.RanksOBJECT import Staff_Info,Ranks
from MyProgram.Upload import imges,upload_image
import os


WorkBook = xl.load_workbook('StaffInfo/StaffInfo.xlsx')
Sheet = WorkBook['Sheet2']

for row in range(2 , Sheet.max_row+1):
    ID = Sheet.cell(row, 1).value
    Name = Sheet.cell(row,2).value
    Rank = Sheet.cell(row, 3).value
    Instgram = Sheet.cell(row, 4).value
    Telegram = Sheet.cell(row, 5).value
    Steam = Sheet.cell(row, 6).value

    if os.path.isfile('StaffCards/{}_{}.png'.format(Name,Rank)):
        print("{}'s Card is already exist".format(Name))
        continue

    print(int(((row - 1) / (Sheet.max_row + 1)) * 100), "%")

    NC_Ranks = Ranks(ID,Name, Rank, Instgram, Telegram, Steam)



    if NC_Ranks.Staff_Rank.upper() == "FOUNDER":
        NC_Ranks.FounderF()

    elif NC_Ranks.Staff_Rank.upper() == "HEADADMIN":
        NC_Ranks.HeadAdminF()


    elif NC_Ranks.Staff_Rank.upper() == "SUPERADMIN":
        NC_Ranks.SuperAdminF()


    elif NC_Ranks.Staff_Rank.upper() == "ADMIN":
        NC_Ranks.AdminF()


    elif NC_Ranks.Staff_Rank.upper() == "ACT":
        NC_Ranks.ACTF()



exit("Done")
